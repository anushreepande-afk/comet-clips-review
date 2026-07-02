#!/usr/bin/env python3
"""Generate clips_data.json from comet_clips.xlsx. Run directly or called by rebuild.command."""
import json
import os
import re
import sys

import openpyxl

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "comet_clips.xlsx")
OUT_PATH  = os.path.join(os.path.dirname(__file__), "clips_data.json")

KEY_COLS = {
    "clip_id":                        "clip_id",
    "clip_drive_link":                "clip_drive_link",
    "Genre CMS":                      "genre_cms",
    "description":                    "description",
    "score":                          "score",
    "tier":                           "tier",
    "full_watch_probability_percent": "watch_prob",
}


def extract_file_id(url: str) -> str:
    if not url:
        return ""
    match = re.search(r"/file/d/([a-zA-Z0-9_-]+)", str(url))
    return match.group(1) if match else ""


def parse_tab_name(name: str):
    parts = name.rsplit("_", 1)
    return (parts[0], parts[1]) if len(parts) == 2 else ("", "")


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    clips = []

    for sheet_name in wb.sheetnames:
        content_id, clip_type = parse_tab_name(sheet_name)
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        col_idx = {}
        for col_name, field in KEY_COLS.items():
            if col_name in headers:
                col_idx[field] = headers.index(col_name)

        for row in ws.iter_rows(min_row=2, values_only=True):
            clip_id_idx = col_idx.get("clip_id", 0)
            if not row[clip_id_idx]:
                continue
            clip = {"content_id": content_id, "clip_type": clip_type}
            for field, idx in col_idx.items():
                val = row[idx]
                clip[field] = val if val is not None else ""
            url = clip.pop("clip_drive_link", "")
            clip["drive_file_id"] = extract_file_id(str(url) if url else "")
            clips.append(clip)

    with open(OUT_PATH, "w") as f:
        json.dump(clips, f, indent=2)
    print(f"Written {len(clips)} clips to {OUT_PATH}")


if __name__ == "__main__":
    main()
