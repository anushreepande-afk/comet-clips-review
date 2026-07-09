import os
import sys
from io import BytesIO

from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from excel_export import (
    build_individual_ratings_workbook,
    build_rating_export_workbook,
    update_workbook_with_rating_summary,
)


def test_build_rating_export_workbook():
    clips = [{
        "content_id": "1260029222",
        "content_name": "Example Movie",
        "clip_type": "momenttype_pro",
        "clip_id": "clip1",
        "clip_drive_link": "https://drive.google.com/file/d/abc/view",
        "genre_cms": "Drama",
        "description": "A key moment.",
    }]
    summary = {"1260029222::momenttype::pro::clip1": {"accept_count": 2, "reject_count": 1, "total": 3, "acceptance_rate": 0.6667}}
    ratings = [{
        "unique_clip_key": "1260029222::momenttype::pro::clip1",
        "content_id": "1260029222",
        "clip_type": "momenttype_pro",
        "clip_id": "clip1",
        "reviewer_email": "reviewer@jiostar.com",
        "score": 1,
        "submitted_at": "2026-07-09T00:00:00Z",
    }]

    data = build_rating_export_workbook(clips, summary, ratings)
    wb = load_workbook(BytesIO(data))
    ws = wb["Example Movie V3"]

    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    assert "Accept Count" in headers
    assert values[headers.index("unique_clip_key")] == "1260029222::momenttype::pro::clip1"
    assert values[headers.index("Accept Count")] == 2
    assert values[headers.index("Reject Count")] == 1
    assert values[headers.index("Total Decisions")] == 3
    assert "Individual Ratings" in wb.sheetnames

    detail_ws = wb["Individual Ratings"]
    detail_headers = [cell.value for cell in detail_ws[1]]
    detail_values = [cell.value for cell in detail_ws[2]]
    assert detail_values[detail_headers.index("reviewer_email")] == "reviewer@jiostar.com"
    assert detail_values[detail_headers.index("decision")] == "Accept"
    assert detail_values[detail_headers.index("clip_drive_link")] == "https://drive.google.com/file/d/abc/view"


def test_build_individual_ratings_workbook():
    clips = [{
        "content_id": "1260029222",
        "content_name": "Example Movie",
        "clip_type": "momenttype_pro",
        "clip_id": "clip1",
        "clip_drive_link": "https://drive.google.com/file/d/abc/view",
    }]
    ratings = [{
        "unique_clip_key": "1260029222::momenttype::pro::clip1",
        "content_id": "1260029222",
        "clip_type": "momenttype_pro",
        "clip_id": "clip1",
        "reviewer_email": "reviewer@jiostar.com",
        "score": 0,
        "submitted_at": "2026-07-09T00:00:00Z",
    }]

    data = build_individual_ratings_workbook(clips, ratings)
    wb = load_workbook(BytesIO(data))

    assert wb.sheetnames == ["Individual Ratings"]
    ws = wb["Individual Ratings"]
    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    assert values[headers.index("reviewer_email")] == "reviewer@jiostar.com"
    assert values[headers.index("decision")] == "Reject"
    assert values[headers.index("clip_drive_link")] == "https://drive.google.com/file/d/abc/view"


def test_update_workbook_with_rating_summary(tmp_path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Example V3"
    ws.append(["content_id", "content_name", "output_set", "clip_id"])
    ws.append(["1260029222", "Example Movie", "V3", "clip1"])
    wb.save(source)

    summary = {"1260029222::momenttype::pro::clip1": {"accept_count": 3, "reject_count": 1, "total": 4, "acceptance_rate": 0.75}}
    ratings = [{
        "unique_clip_key": "1260029222::momenttype::pro::clip1",
        "content_id": "1260029222",
        "clip_type": "momenttype_pro",
        "clip_id": "clip1",
        "reviewer_email": "reviewer@jiostar.com",
        "score": 1,
        "submitted_at": "2026-07-09T00:00:00Z",
    }]
    updated = update_workbook_with_rating_summary(source, output, summary, ratings)

    assert updated == 1
    out_wb = load_workbook(output)
    out_ws = out_wb.active
    headers = [cell.value for cell in out_ws[1]]
    values = [cell.value for cell in out_ws[2]]
    assert values[headers.index("Unique Clip Key")] == "1260029222::momenttype::pro::clip1"
    assert values[headers.index("Accept Count")] == 3
    assert values[headers.index("Reject Count")] == 1
    assert values[headers.index("Total Decisions")] == 4
    assert "Individual Ratings" in out_wb.sheetnames

    detail_ws = out_wb["Individual Ratings"]
    detail_headers = [cell.value for cell in detail_ws[1]]
    detail_values = [cell.value for cell in detail_ws[2]]
    assert detail_values[detail_headers.index("reviewer_email")] == "reviewer@jiostar.com"
    assert detail_values[detail_headers.index("decision")] == "Accept"
