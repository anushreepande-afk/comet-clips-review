from __future__ import annotations

import re
from collections import defaultdict
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from db import build_unique_clip_key

OUTPUT_SET_ORDER = [
    "cliffhanger_pro",
    "cliffhanger_flash",
    "momenttype_pro",
    "momenttype_flash",
]

OUTPUT_SET_LABELS: Dict[str, str] = {
    "cliffhanger_pro": "V1",
    "cliffhanger_flash": "V2",
    "momenttype_pro": "V3",
    "momenttype_flash": "V4",
}

LABEL_TO_CLIP_TYPE = {label.lower(): clip_type for clip_type, label in OUTPUT_SET_LABELS.items()}
LABEL_TO_CLIP_TYPE.update({
    "alpha": "cliffhanger_pro",
    "beta": "cliffhanger_flash",
    "gamma": "momenttype_pro",
    "delta": "momenttype_flash",
})

EXPORT_HEADERS = [
    "content_id",
    "content_name",
    "output_set",
    "clip_id",
    "unique_clip_key",
    "clip_drive_link",
    "Genre CMS",
    "description",
    "Accept Count",
    "Reject Count",
    "Total Decisions",
    "Acceptance Rate",
    "Rejection Rating Count",
    "Average Rejection Rating",
    "Rejection Feedback Count",
]

RATING_EXPORT_COLUMNS = [
    "Unique Clip Key",
    "Accept Count",
    "Reject Count",
    "Total Decisions",
    "Acceptance Rate",
    "Rejection Rating Count",
    "Average Rejection Rating",
    "Rejection Feedback Count",
]

INDIVIDUAL_RATING_HEADERS = [
    "unique_clip_key",
    "content_id",
    "content_name",
    "output_set",
    "clip_id",
    "clip_drive_link",
    "reviewer_email",
    "decision",
    "rejection_rating",
    "rejection_feedback",
    "submitted_at",
]


def _natural_clip_number(clip_id: str) -> int:
    match = re.search(r"(\d+)$", str(clip_id))
    return int(match.group(1)) if match else 999999


def _safe_sheet_title(base: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", str(base)).strip() or "Sheet"
    cleaned = re.sub(r"\s+", " ", cleaned)
    title = cleaned[:31]
    suffix = 2
    while title in used:
        tail = f" {suffix}"
        title = f"{cleaned[:31 - len(tail)]}{tail}"
        suffix += 1
    used.add(title)
    return title


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 14,
        "B": 24,
        "C": 12,
        "D": 12,
        "E": 38,
        "F": 48,
        "G": 16,
        "H": 52,
        "I": 16,
        "J": 14,
        "K": 16,
        "L": 16,
        "M": 22,
        "N": 24,
        "O": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_individual_ratings_sheet(ws) -> None:
    _style_sheet(ws)
    widths = {
        "A": 38,
        "B": 14,
        "C": 24,
        "D": 12,
        "E": 12,
        "F": 48,
        "G": 30,
        "H": 14,
        "I": 18,
        "J": 42,
        "K": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _summary_values(
    rating_summary: Dict[str, Dict],
    unique_clip_key: str,
) -> tuple[int, int, int, Optional[float], int, Optional[float], int]:
    summary = rating_summary.get(unique_clip_key) or {}
    accepts = int(summary.get("accept_count") or summary.get("accepts") or 0)
    rejects = int(summary.get("reject_count") or summary.get("rejects") or 0)
    total = int(summary.get("total") or summary.get("count") or accepts + rejects)
    rate = summary.get("acceptance_rate")
    if rate is None and total:
        rate = round(accepts / total, 4)
    rejection_rating_count = int(summary.get("rejection_rating_count") or 0)
    avg_rejection_rating = summary.get("avg_rejection_rating")
    rejection_feedback_count = int(summary.get("rejection_feedback_count") or 0)
    return accepts, rejects, total, rate, rejection_rating_count, avg_rejection_rating, rejection_feedback_count


def _decision_from_score(score: object) -> str:
    try:
        return "Accept" if int(score) == 1 else "Reject"
    except Exception:
        return ""


def _clip_meta_by_key(clips: Iterable[Dict]) -> Dict[str, Dict]:
    meta = {}
    for clip in clips:
        unique_clip_key = build_unique_clip_key(clip["content_id"], clip["clip_type"], clip["clip_id"])
        meta[unique_clip_key] = {
            "content_id": clip["content_id"],
            "content_name": clip.get("content_name", ""),
            "output_set": OUTPUT_SET_LABELS.get(clip["clip_type"], clip["clip_type"]),
            "clip_id": clip["clip_id"],
            "clip_drive_link": clip.get("clip_drive_link", ""),
        }
    return meta


def _append_individual_ratings_sheet(wb, clip_meta: Dict[str, Dict], ratings: Iterable[Dict]) -> None:
    ws = wb.create_sheet("Individual Ratings")
    ws.append(INDIVIDUAL_RATING_HEADERS)

    sorted_ratings = sorted(
        ratings,
        key=lambda row: (
            str(row.get("unique_clip_key") or ""),
            str(row.get("reviewer_email") or ""),
            str(row.get("submitted_at") or ""),
        ),
    )

    for rating in sorted_ratings:
        unique_clip_key = rating.get("unique_clip_key") or build_unique_clip_key(
            rating["content_id"],
            rating["clip_type"],
            rating["clip_id"],
        )
        meta = clip_meta.get(unique_clip_key, {})
        ws.append([
            unique_clip_key,
            rating.get("content_id") or meta.get("content_id", ""),
            meta.get("content_name", ""),
            meta.get("output_set", rating.get("clip_type", "")),
            rating.get("clip_id") or meta.get("clip_id", ""),
            meta.get("clip_drive_link", ""),
            rating.get("reviewer_email", ""),
            rating.get("decision") or _decision_from_score(rating.get("score")),
            rating.get("rejection_rating", "") if _decision_from_score(rating.get("score")) == "Reject" else "",
            rating.get("feedback_text", "") if _decision_from_score(rating.get("score")) == "Reject" else "",
            rating.get("submitted_at", ""),
        ])
        link_cell = ws.cell(row=ws.max_row, column=6)
        if link_cell.value:
            link_cell.hyperlink = link_cell.value
            link_cell.style = "Hyperlink"

    _style_individual_ratings_sheet(ws)


def build_rating_export_workbook(
    clips: Iterable[Dict],
    rating_summary: Dict[str, Dict],
    ratings: Optional[Iterable[Dict]] = None,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    clips = list(clips)

    grouped: Dict[tuple[str, str], List[Dict]] = defaultdict(list)
    for clip in clips:
        grouped[(clip["content_id"], clip["clip_type"])].append(clip)

    used_titles: set[str] = set()
    ordered_keys = sorted(
        grouped,
        key=lambda key: (key[0], OUTPUT_SET_ORDER.index(key[1]) if key[1] in OUTPUT_SET_ORDER else 99),
    )

    for content_id, clip_type in ordered_keys:
        rows = sorted(grouped[(content_id, clip_type)], key=lambda c: _natural_clip_number(c["clip_id"]))
        first = rows[0]
        output_label = OUTPUT_SET_LABELS.get(clip_type, clip_type)
        sheet_title = _safe_sheet_title(f"{first.get('content_name', content_id)} {output_label}", used_titles)
        ws = wb.create_sheet(sheet_title)
        ws.append(EXPORT_HEADERS)

        for clip in rows:
            unique_clip_key = build_unique_clip_key(clip["content_id"], clip["clip_type"], clip["clip_id"])
            (
                accepts,
                rejects,
                total,
                acceptance_rate,
                rejection_rating_count,
                avg_rejection_rating,
                rejection_feedback_count,
            ) = _summary_values(rating_summary, unique_clip_key)
            ws.append([
                clip["content_id"],
                clip.get("content_name", ""),
                output_label,
                clip["clip_id"],
                unique_clip_key,
                clip.get("clip_drive_link", ""),
                clip.get("genre_cms", ""),
                clip.get("description", ""),
                accepts,
                rejects,
                total,
                acceptance_rate,
                rejection_rating_count,
                avg_rejection_rating,
                rejection_feedback_count,
            ])
            link_cell = ws.cell(row=ws.max_row, column=6)
            if link_cell.value:
                link_cell.hyperlink = link_cell.value
                link_cell.style = "Hyperlink"

        _style_sheet(ws)

    if ratings is not None:
        _append_individual_ratings_sheet(wb, _clip_meta_by_key(clips), ratings)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_individual_ratings_workbook(clips: Iterable[Dict], ratings: Iterable[Dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _append_individual_ratings_sheet(wb, _clip_meta_by_key(list(clips)), ratings)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _header_map(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        normalized = _normalize_header(cell.value)
        if normalized:
            headers[normalized] = idx
    return headers


def _ensure_column(ws, header_map: Dict[str, int], title: str) -> int:
    normalized = _normalize_header(title)
    if normalized in header_map:
        return header_map[normalized]
    col_idx = ws.max_column + 1
    ws.cell(row=1, column=col_idx, value=title)
    header_map[normalized] = col_idx
    return col_idx


def _cell_value(row, header_map: Dict[str, int], *names: str) -> str:
    for name in names:
        idx = header_map.get(_normalize_header(name))
        if idx:
            value = row[idx - 1].value
            if value is not None:
                return str(value).strip()
    return ""


def infer_clip_type(sheet_name: str, row, header_map: Dict[str, int]) -> Optional[str]:
    explicit = _cell_value(row, header_map, "clip_type")
    if explicit:
        return explicit

    output_value = _cell_value(row, header_map, "output_set", "output_label")
    mapped = LABEL_TO_CLIP_TYPE.get(output_value.lower())
    if mapped:
        return mapped

    text = sheet_name.lower().replace("-", " ").replace("_", " ")
    for label, clip_type in LABEL_TO_CLIP_TYPE.items():
        if label in text:
            return clip_type
    if "cliffhanger" in text and "pro" in text:
        return "cliffhanger_pro"
    if "cliffhanger" in text and "flash" in text:
        return "cliffhanger_flash"
    if ("moment type" in text or "momenttype" in text) and "pro" in text:
        return "momenttype_pro"
    if ("moment type" in text or "momenttype" in text) and "flash" in text:
        return "momenttype_flash"
    return None


def update_workbook_with_rating_summary(
    input_path: str | Path,
    output_path: str | Path,
    rating_summary: Dict[str, Dict],
    ratings: Optional[Iterable[Dict]] = None,
) -> int:
    wb = load_workbook(input_path)
    updated_rows = 0
    clip_meta: Dict[str, Dict] = {}

    for ws in wb.worksheets:
        if ws.title == "Individual Ratings":
            continue
        if ws.max_row < 2:
            continue
        headers = _header_map(ws)
        content_col = headers.get("contentid")
        clip_col = headers.get("clipid")
        if not content_col or not clip_col:
            continue

        unique_key_col = _ensure_column(ws, headers, RATING_EXPORT_COLUMNS[0])
        accept_col = _ensure_column(ws, headers, RATING_EXPORT_COLUMNS[1])
        reject_col = _ensure_column(ws, headers, RATING_EXPORT_COLUMNS[2])
        total_col = _ensure_column(ws, headers, RATING_EXPORT_COLUMNS[3])
        rate_col = _ensure_column(ws, headers, RATING_EXPORT_COLUMNS[4])
        rejection_rating_count_col = _ensure_column(ws, headers, RATING_EXPORT_COLUMNS[5])
        avg_rejection_rating_col = _ensure_column(ws, headers, RATING_EXPORT_COLUMNS[6])
        rejection_feedback_count_col = _ensure_column(ws, headers, RATING_EXPORT_COLUMNS[7])

        for header_cell in ws[1]:
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            content_id = str(row[content_col - 1].value or "").strip()
            clip_id = str(row[clip_col - 1].value or "").strip()
            if not content_id or not clip_id:
                continue
            clip_type = infer_clip_type(ws.title, row, headers)
            if not clip_type:
                continue
            unique_clip_key = build_unique_clip_key(content_id, clip_type, clip_id)
            (
                accepts,
                rejects,
                total,
                acceptance_rate,
                rejection_rating_count,
                avg_rejection_rating,
                rejection_feedback_count,
            ) = _summary_values(rating_summary, unique_clip_key)
            output_set = _cell_value(row, headers, "output_set", "output_label") or OUTPUT_SET_LABELS.get(clip_type, clip_type)
            clip_meta[unique_clip_key] = {
                "content_id": content_id,
                "content_name": _cell_value(row, headers, "content_name"),
                "output_set": output_set,
                "clip_id": clip_id,
                "clip_drive_link": _cell_value(row, headers, "clip_drive_link", "clip link", "clip drive link"),
            }

            ws.cell(row=row[0].row, column=unique_key_col, value=unique_clip_key)
            ws.cell(row=row[0].row, column=accept_col, value=accepts)
            ws.cell(row=row[0].row, column=reject_col, value=rejects)
            ws.cell(row=row[0].row, column=total_col, value=total)
            ws.cell(row=row[0].row, column=rate_col, value=acceptance_rate)
            ws.cell(row=row[0].row, column=rejection_rating_count_col, value=rejection_rating_count)
            ws.cell(row=row[0].row, column=avg_rejection_rating_col, value=avg_rejection_rating)
            ws.cell(row=row[0].row, column=rejection_feedback_count_col, value=rejection_feedback_count)
            updated_rows += 1

        for col_idx in [
            unique_key_col,
            accept_col,
            reject_col,
            total_col,
            rate_col,
            rejection_rating_count_col,
            avg_rejection_rating_col,
            rejection_feedback_count_col,
        ]:
            ws.column_dimensions[get_column_letter(col_idx)].width = 18 if col_idx != unique_key_col else 38

    if ratings is not None:
        if "Individual Ratings" in wb.sheetnames:
            del wb["Individual Ratings"]
        _append_individual_ratings_sheet(wb, clip_meta, ratings)

    wb.save(output_path)
    return updated_rows
